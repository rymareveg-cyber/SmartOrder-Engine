#!/usr/bin/env python3
"""
Data Exporter - модуль экспорта данных в Excel, CSV и PDF.

Предоставляет функции для экспорта заказов и статистики в различные форматы.
"""

import os
import csv
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from pathlib import Path as PathLib
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from src.services.order_service import OrderService, Order
from src.utils.logger import get_logger

logger = get_logger(__name__)

EXPORT_DIR = PathLib(__file__).parent.parent / ".tmp" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

CYRILLIC_FONT = "Arial"
CYRILLIC_FONT_BOLD = "Arial-Bold"

try:
    # Попытка зарегистрировать системные шрифты
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont("Helvetica"))
    CYRILLIC_FONT = "Helvetica"
    CYRILLIC_FONT_BOLD = "Helvetica-Bold"
except:
    try:
        # Попытка загрузить Arial из Windows
        if os.name == 'nt':
            font_path = "C:/Windows/Fonts/arial.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("Arial", font_path))
                CYRILLIC_FONT = "Arial"
                font_path_bold = "C:/Windows/Fonts/arialbd.ttf"
                if os.path.exists(font_path_bold):
                    pdfmetrics.registerFont(TTFont("Arial-Bold", font_path_bold))
                    CYRILLIC_FONT_BOLD = "Arial-Bold"
    except Exception as e:
        logger.warning(f"Could not register Cyrillic fonts: {e}")


class DataExporter:
    """Класс для экспорта данных в различные форматы."""
    
    @staticmethod
    def export_orders_to_excel(
        orders: List[Order],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт заказов в Excel.
        
        Args:
            orders: Список заказов
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному Excel файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"orders_export_{timestamp}.xlsx"
        
        filepath = EXPORT_DIR / filename
        
        # Создание рабочей книги
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        # Стили заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = [
            "Номер заказа", "Статус", "Канал", "Клиент", "Телефон", 
            "Адрес", "Сумма товаров", "Доставка", "Итого", 
            "Дата создания", "Дата оплаты", "Трек-номер", "ID транзакции"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        for row_num, order in enumerate(orders, 2):
            # Расчет суммы товаров
            items_total = sum(item.total for item in order.items)
            
            row_data = [
                order.order_number,
                order.status,
                order.channel,
                order.customer_name or "",
                order.customer_phone or "",
                order.customer_address or "",
                items_total,
                order.delivery_cost,
                order.total_amount,
                order.created_at,
                order.paid_at or "",
                order.tracking_number or "",
                order.transaction_id or ""
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                if col_num in [7, 8, 9]:  # Числовые колонки
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
        
        # Автоматическая ширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение
        wb.save(str(filepath))
        logger.info(f"Exported {len(orders)} orders to Excel: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def export_orders_to_csv(
        orders: List[Order],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт заказов в CSV.
        
        Args:
            orders: Список заказов
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному CSV файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"orders_export_{timestamp}.csv"
        
        filepath = EXPORT_DIR / filename
        
        # Заголовки
        headers = [
            "Номер заказа", "Статус", "Канал", "Клиент", "Телефон", 
            "Адрес", "Сумма товаров", "Доставка", "Итого", 
            "Дата создания", "Дата оплаты", "Трек-номер", "ID транзакции"
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)
            
            for order in orders:
                items_total = sum(item.total for item in order.items)
                
                row = [
                    order.order_number,
                    order.status,
                    order.channel,
                    order.customer_name or "",
                    order.customer_phone or "",
                    order.customer_address or "",
                    items_total,
                    order.delivery_cost,
                    order.total_amount,
                    order.created_at,
                    order.paid_at or "",
                    order.tracking_number or "",
                    order.transaction_id or ""
                ]
                writer.writerow(row)
        
        logger.info(f"Exported {len(orders)} orders to CSV: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def export_stats_to_pdf(
        stats: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт статистики в PDF.
        
        Args:
            stats: Словарь со статистикой
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному PDF файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"stats_export_{timestamp}.pdf"
        
        filepath = EXPORT_DIR / filename
        
        # Создание PDF документа
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # Стили
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=CYRILLIC_FONT,
            fontSize=16,
            textColor=colors.HexColor('#000000'),
            alignment=1,  # Center
            spaceAfter=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=CYRILLIC_FONT,
            fontSize=10,
            leading=12
        )
        
        bold_style = ParagraphStyle(
            'CustomBold',
            parent=styles['Normal'],
            fontName=CYRILLIC_FONT_BOLD,
            fontSize=10,
            leading=12
        )
        
        # Содержимое документа
        story = []
        
        # Заголовок
        story.append(Paragraph("Отчёт по статистике SmartOrder Engine", title_style))
        story.append(Spacer(1, 10*mm))
        
        # Дата генерации
        story.append(Paragraph(
            f"Дата генерации: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}",
            normal_style
        ))
        story.append(Spacer(1, 5*mm))
        
        # Основные метрики
        story.append(Paragraph("Основные метрики", bold_style))
        story.append(Spacer(1, 3*mm))
        
        metrics_data = [
            ["Метрика", "Значение"],
            ["Выручка сегодня", f"{stats.get('revenue_today', 0):,.2f} ₽"],
            ["Выручка за неделю", f"{stats.get('revenue_week', 0):,.2f} ₽"],
            ["Выручка за месяц", f"{stats.get('revenue_month', 0):,.2f} ₽"],
            ["Заказов сегодня", str(stats.get('orders_today', 0))],
            ["Заказов за неделю", str(stats.get('orders_week', 0))],
            ["Заказов за месяц", str(stats.get('orders_month', 0))],
            ["Конверсия", f"{stats.get('conversion_rate', 0):.2f}%"],
            ["Средний чек", f"{stats.get('average_check', 0):,.2f} ₽"],
        ]
        
        table = Table(metrics_data, colWidths=[100*mm, 70*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), CYRILLIC_FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), CYRILLIC_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 10*mm))
        
        # Топ товаров
        if stats.get('top_products'):
            story.append(Paragraph("Топ товаров", bold_style))
            story.append(Spacer(1, 3*mm))
            
            products_data = [["№", "Артикул", "Название", "Количество", "Выручка"]]
            for idx, product in enumerate(stats['top_products'][:10], 1):
                products_data.append([
                    str(idx),
                    product.get('articul', ''),
                    product.get('name', ''),
                    str(product.get('quantity', 0)),
                    f"{product.get('revenue', 0):,.2f} ₽"
                ])
            
            products_table = Table(products_data, colWidths=[10*mm, 30*mm, 60*mm, 20*mm, 30*mm])
            products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), CYRILLIC_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), CYRILLIC_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(products_table)
        
        # Генерация PDF
        doc.build(story)
        logger.info(f"Exported statistics to PDF: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def export_analytics_to_pdf(
        analytics: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт аналитики в PDF.
        
        Args:
            analytics: Словарь с аналитикой
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному PDF файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"analytics_export_{timestamp}.pdf"
        
        filepath = EXPORT_DIR / filename
        
        # Создание PDF документа
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # Стили
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=CYRILLIC_FONT,
            fontSize=16,
            textColor=colors.HexColor('#000000'),
            alignment=1,  # Center
            spaceAfter=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=CYRILLIC_FONT,
            fontSize=10,
            leading=12
        )
        
        bold_style = ParagraphStyle(
            'CustomBold',
            parent=styles['Normal'],
            fontName=CYRILLIC_FONT_BOLD,
            fontSize=10,
            leading=12
        )
        
        # Содержимое документа
        story = []
        
        # Заголовок
        story.append(Paragraph("Детальная аналитика SmartOrder Engine", title_style))
        story.append(Spacer(1, 10*mm))
        
        # Дата генерации
        story.append(Paragraph(
            f"Дата генерации: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}",
            normal_style
        ))
        story.append(Spacer(1, 5*mm))
        
        # Анализ по каналам
        if analytics.get('channel_analysis'):
            story.append(Paragraph("Анализ по каналам", bold_style))
            story.append(Spacer(1, 3*mm))
            
            channel_data = [["Канал", "Заказов", "Выручка", "Средний чек"]]
            for channel, data in analytics['channel_analysis'].items():
                channel_name = {
                    'telegram': 'Telegram',
                    'yandex_mail': 'Яндекс.Почта',
                    'yandex_forms': 'Яндекс.Формы'
                }.get(channel, channel)
                channel_data.append([
                    channel_name,
                    str(data.get('orders_count', 0)),
                    f"{data.get('revenue', 0):,.2f} ₽",
                    f"{data.get('avg_order_value', 0):,.2f} ₽"
                ])
            
            channel_table = Table(channel_data, colWidths=[40*mm, 30*mm, 40*mm, 40*mm])
            channel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), CYRILLIC_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), CYRILLIC_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(channel_table)
            story.append(Spacer(1, 10*mm))
        
        # Воронка продаж
        if analytics.get('sales_funnel'):
            story.append(Paragraph("Воронка продаж", bold_style))
            story.append(Spacer(1, 3*mm))
            
            funnel_data = [["Этап", "Количество"]]
            status_names = {
                'new': 'Новые',
                'validated': 'Валидированные',
                'invoice_created': 'Счета созданы',
                'paid': 'Оплаченные',
                'shipped': 'Отправленные',
                'cancelled': 'Отмененные'
            }
            for status, count in analytics['sales_funnel'].items():
                funnel_data.append([
                    status_names.get(status, status),
                    str(count)
                ])
            
            funnel_table = Table(funnel_data, colWidths=[100*mm, 70*mm])
            funnel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), CYRILLIC_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), CYRILLIC_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(funnel_table)
            story.append(Spacer(1, 10*mm))
        
        # Дополнительные метрики
        if analytics.get('metrics'):
            story.append(Paragraph("Дополнительные метрики", bold_style))
            story.append(Spacer(1, 3*mm))
            
            metrics = analytics['metrics']
            metrics_data = [["Метрика", "Значение"]]
            if metrics.get('avg_processing_hours'):
                metrics_data.append(["Среднее время обработки", f"{metrics['avg_processing_hours']:.1f} ч."])
            if metrics.get('avg_delivery_hours'):
                metrics_data.append(["Среднее время доставки", f"{metrics['avg_delivery_hours']:.1f} ч."])
            if metrics.get('avg_delivery_cost'):
                metrics_data.append(["Средняя стоимость доставки", f"{metrics['avg_delivery_cost']:,.2f} ₽"])
            if metrics.get('orders_with_delivery'):
                metrics_data.append(["Заказов с доставкой", str(metrics['orders_with_delivery'])])
            
            if len(metrics_data) > 1:
                metrics_table = Table(metrics_data, colWidths=[100*mm, 70*mm])
                metrics_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), CYRILLIC_FONT_BOLD),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), CYRILLIC_FONT),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                story.append(metrics_table)
        
        # Генерация PDF
        doc.build(story)
        logger.info(f"Exported analytics to PDF: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def export_catalog_to_excel(
        products: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт каталога в Excel.
        
        Args:
            products: Список товаров
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному Excel файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"catalog_export_{timestamp}.xlsx"
        
        filepath = EXPORT_DIR / filename
        
        # Создание рабочей книги
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"
        
        # Стили заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ["Артикул", "Название", "Цена", "Остаток", "Обновлено"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        for row_num, product in enumerate(products, 2):
            row_data = [
                product.get('articul', ''),
                product.get('name', ''),
                product.get('price', 0),
                product.get('stock', 0),
                product.get('updated_at', '') or ''
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                if col_num == 3:  # Цена
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 4:  # Остаток
                    cell.alignment = Alignment(horizontal="right")
        
        # Автоматическая ширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение
        wb.save(str(filepath))
        logger.info(f"Exported {len(products)} products to Excel: {filepath}")
        
        return str(filepath)
    
    @staticmethod
    def export_catalog_to_csv(
        products: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        Экспорт каталога в CSV.
        
        Args:
            products: Список товаров
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному CSV файлу
        """
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"catalog_export_{timestamp}.csv"
        
        filepath = EXPORT_DIR / filename
        
        # Заголовки
        headers = ["Артикул", "Название", "Цена", "Остаток", "Обновлено"]
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)
            
            for product in products:
                row = [
                    product.get('articul', ''),
                    product.get('name', ''),
                    product.get('price', 0),
                    product.get('stock', 0),
                    product.get('updated_at', '') or ''
                ]
                writer.writerow(row)
        
        logger.info(f"Exported {len(products)} products to CSV: {filepath}")
        
        return str(filepath)


if __name__ == "__main__":
    # Тестирование экспорта
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python data_exporter.py <format> [order_id]")
        print("Formats: excel, csv, pdf")
        sys.exit(1)
    
    format_type = sys.argv[1]
    
    if format_type in ['excel', 'csv']:
        # Экспорт заказов
        orders_data = OrderService.list_orders(page_size=100)
        orders = orders_data['items']
        
        if format_type == 'excel':
            filepath = DataExporter.export_orders_to_excel(orders)
        else:
            filepath = DataExporter.export_orders_to_csv(orders)
        
        print(f"Exported {len(orders)} orders to {format_type.upper()}: {filepath}")
    elif format_type == 'pdf':
        # Экспорт статистики (требует передачи stats извне)
        print("PDF export requires stats dictionary. Use via API endpoint instead.")
        sys.exit(1)
    else:
        print(f"Unknown format: {format_type}")
        sys.exit(1)
